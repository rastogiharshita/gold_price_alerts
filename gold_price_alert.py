import logging
from datetime import date
from pathlib import Path
from openpyxl import Workbook, load_workbook
import requests
import urllib3
import pandas as pd

import telegram_alert as tg

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
log_file = Path(__file__).with_suffix(".log")
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(asctime)s: %(message)s",
                    filename=log_file, encoding="utf-8")
logger = logging.getLogger(__name__)

IBJA_API_URL = "https://ibja-api.vercel.app/latest"
TIMEOUT_SECONDS = 60
FILEPATH = Path(__file__).with_name("gold_price_tracking.xlsx")
SHEET_NAME = "gold_price"
HEADERS = ("Date", "Gold Price (24K)")


def get_current_gold_price() -> float:
    try:
        logger.info("Fetching gold price from IBJA API")
        response = requests.get(url=IBJA_API_URL, timeout=TIMEOUT_SECONDS, verify=False)
        response.raise_for_status()
        price = response.json().get("lblGold999_AM")
        return float(price)
    except Exception as exc:
        logger.error("Failed to fetch gold price: %s", exc)

def get_or_create_sheet(workbook: Workbook):
    if SHEET_NAME in workbook.sheetnames:
        sheet = workbook[SHEET_NAME]
    else:
        sheet = workbook.active if workbook.active.max_row == 1 and workbook.active.max_column == 1 and workbook.active["A1"].value is None else workbook.create_sheet(SHEET_NAME)
        sheet.title = SHEET_NAME

    if sheet["A1"].value != HEADERS[0] or sheet["B1"].value != HEADERS[1]:
        sheet["A1"] = HEADERS[0]
        sheet["B1"] = HEADERS[1]

    return sheet


def append_to_track_sheet(price: float) -> None:
    workbook = load_workbook(FILEPATH) if FILEPATH.exists() else Workbook()
    sheet = get_or_create_sheet(workbook)
    sheet.append([date.today().isoformat(), price])
    workbook.save(FILEPATH)
    logger.info("Saved gold price to %s", FILEPATH)


def calculate_feasible_price_range() -> (float, float):
    price_df = pd.read_excel(FILEPATH)
    price_df['Date'] = pd.to_datetime(price_df['Date'])
    price_df = price_df.groupby('Date')['Gold Price (24K)'].mean().reset_index()
    current_date = pd.Timestamp.today().normalize()
    six_month_prev_date = current_date - pd.DateOffset(months=6)
    filtered_df = price_df[(price_df['Date'] >= six_month_prev_date) & (price_df['Date'] <= current_date)]
    mean_price = filtered_df['Gold Price (24K)'].mean()
    return mean_price, mean_price * 1.1

def generate_alert(min_price, max_price, current_price):
    logger.info("Generating gold price alert")
    if current_price <= max_price:
        tg.send_message_to_telegram(f"Min: {min_price}, Max: {max_price}, Current Price: {current_price}")
        tg.send_message_to_telegram("Good Time to buy!")
    elif current_price <= min_price:
        tg.send_message_to_telegram(f"Price DROP !!! Min: {min_price}, Current Price: {current_price}")
    else:
        tg.send_message_to_telegram("Too Expensive !! Next time .")

if __name__ == "__main__":
    price = get_current_gold_price()
    if price:
        append_to_track_sheet(price=price)
        min_price, max_price = calculate_feasible_price_range()
        generate_alert(min_price, max_price, price)
    else:
        tg.send_message_to_telegram("Market closed today !")
